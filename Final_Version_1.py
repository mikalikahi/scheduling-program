from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import Menu
import smtplib
import openpyxl
import datetime
from tkinter import Spinbox
#from time import  sleep
import time
today = datetime.datetime.now()

#=========================================================================================================================================================
class ToolTip(object):
    def __init__(self, widget):
        self.widget= widget
        self.tipwindow = None
        self.id = None
        self.x = self.y = 0

    def showtip(self, text):
        "Display text in tooltip window"
        self.text = text
        if self.tipwindow or not self.text:
            return
        x, y, _cx, cy = self.widget.bbox("insert")
        x = x + self.widget.winfo_rootx() + 27
        y = y + cy + self.widget.winfo_rooty() + 27
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(1)
        tw.wm_geometry("+%d+%d" % (x, y))

        label = tk.Label(tw, text=self.text, justify=tk.LEFT, background="#ffffe0",
                         relief=tk.SOLID, borderwidth=1, font=("tahoma", "8", "normal"))
        label.pack(ipadx=1)

    def hidetip(self):
        tw = self.tipwindow
        self.tipwindow = None
        if tw:
            tw.destroy()

# ====================================================================================================================================================
def createToolTip(widget, text):
    toolTip = ToolTip(widget)
    def enter(event):
        toolTip.showtip(text)
    def leave(event):
        toolTip.hidetip()
    widget.bind('<Enter>', enter)
    widget.bind('<Leave>', leave)
# ====================================================================================================================================================
class Get_Data_From_Excel:

# Methods for teacher search *************************************************************************************************************************
    def get_teacher_list():
        wb = openpyxl.load_workbook('TLI_August_2017_schedule.xlsx')
        sheetList = wb.sheetnames #wb.get_sheet_names()
        return sheetList
    
    # get date list method
    def get_date_list():
        # get teacher list method
        wb = openpyxl.load_workbook('TLI_August_2017_schedule.xlsx')
        sheetList = wb.sheetnames #wb.get_sheet_names()
        # get date list and time list methods
        sheet = wb[sheetList[0]] #wb.get_sheet_by_name(sheetList[0])
        date_List = []
        for d in range(2, 125):
            date_List.append(sheet.cell(row=d, column=1).value[0:11])
        return date_List
    
    # get time list method
    def get_time_list():
        wb = openpyxl.load_workbook('TLI_August_2017_schedule.xlsx')
        sheetList = wb.sheetnames #wb.get_sheet_names()
        lsh = len(sheetList)
        sheet = wb[sheetList[0]] #wb.get_sheet_by_name(sheetList[0])
        time_List = []
        for t in range(2, 16):
            time_List.append(sheet.cell(row=1, column=t).value)
        return time_List
    
    def get_teacher_emails():
        wb = openpyxl.load_workbook('TLI_August_2017_schedule.xlsx')
        sheetList = wb.sheetnames #wb.get_sheet_names()
        teachers_email_list = []
        for t in range(0, len(sheetList)):
            sheet = wb[sheetList[t]] #wb.get_sheet_by_name(teacher_List[t])
            teachers_email_list.append(sheet.cell(row=2, column=18).value)
        return teachers_email_list

    # method for getting sheet with an arguments for the time and date
    def get_unavailable_teachers(r, c):
        # get not free teachers and get not free teachers' emails
        wb = openpyxl.load_workbook('TLI_August_2017_schedule.xlsx')
        sheetList = wb.sheetnames
        not_free_teachers_list = []
        for t in range(0, len(sheetList)):
            sheet = wb[sheetList[t]] 
            if sheet.cell(row=r, column=c).value != None and sheet.cell(row=r, column=c).value != ' ':
                    not_free_teachers_list.append(sheetList[t])
        return not_free_teachers_list

    # method for getting available teachers' email addresses
    def get_free_teachers_emails(free_teachers_List):
        wb = openpyxl.load_workbook('TLI_August_2017_schedule.xlsx')
        email_list = []
        for teacher in range(0, len(free_teachers_List)):
            sheet = wb[free_teachers_List[teacher]]
            email_list.append(sheet.cell(row=2, column=18).value)
        return email_list

# Methods for emailing schedules to teachers *********************************************************************************************************
    def get_schedule_times():
        wb = openpyxl.load_workbook('TLI_August_2017_schedule.xlsx')
        sheetList = wb.sheetnames
        sheet = wb[sheetList[0]]
        time_List = []
        for i in range(2, 16):
            time_List.append(sheet.cell(row=1, column=i).value)
        return time_List

    def get_schedule_student_names(sheetNumber, schedulingDateRow):
        wb = openpyxl.load_workbook('TLI_August_2017_schedule.xlsx')
        sheetList = wb.sheetnames
        sheet = wb[sheetList[sheetNumber]]
        name_List = []
        for i in range(2, 16):
            if sheet.cell(row=schedulingDateRow, column=i).value == None:
                name_List.append(' ')
            else:
                name_List.append(sheet.cell(row=schedulingDateRow, column=i).value)
        return name_List

# Methods for monthly display of teachers' schedules *************************************************************************************************
    def open_workbook():
        wb = openpyxl.load_workbook('TLI_August_2017_schedule.xlsx')
        return wb
        

    def get_teacher_list():
        wb = openpyxl.load_workbook('TLI_August_2017_schedule.xlsx')
        sheetList = wb.sheetnames #wb.get_sheet_names()
        return sheetList


    def get_current_month_list_and_month_list_values():
        today = datetime.datetime.now()
        month_List = ['January', 'February', 'March', 'April', 'May', 'June',
                     'July', 'August', 'September', 'October', 'November',
                     'December']
        # Figure out months for input beginning with the current month
        this_month = today.month
        month_list_index = this_month - 1
        this_year = today.year
        next_year = this_year + 1

        leap_years = ['2020', '2024', '2028', '2032', '2036', '2040']
        leap_march = 29
        regular_march = 28
        thirty_day_months = ['April', 'June', 'September', 'November']

        current_month_list = []
        current_month_list_values = []

        this_year_month_list = []
        this_year_month_list_values = []

        next_year_month_list = []
        next_year_month_list_values = []

        for i in range(month_list_index, len(month_List)):
                       current_month_list.append(month_List[i])
                       this_year_month_list.append(month_List[i])

        for i in range(0, month_list_index):
                       current_month_list.append(month_List[i])
                       next_year_month_list.append(month_List[i])
                       
        for x in range(0, len(this_year_month_list)):
            if this_year_month_list[x] == 'March':
                if this_year in leap_years:
                    this_year_month_list_values.append(29)
                else:
                    this_year_month_list_values.append(28)

            elif this_year_month_list[x] in thirty_day_months:
                this_year_month_list_values.append(30)
            else:
                this_year_month_list_values.append(31)

        for x in range(0, len(next_year_month_list)):
            if next_year_month_list[x] == 'March':
                if next_year in leap_years:
                    next_year_month_list_values.append(29)
                else:
                    next_year_month_list_values.append(28)

            elif next_year_month_list[x] in thirty_day_months:
                next_year_month_list_values.append(30)
            else:
                next_year_month_list_values.append(31)
        current_month_list_values = this_year_month_list_values + next_year_month_list_values
        return [current_month_list, current_month_list_values]


    def get_start_row_finish_row(month, current_month_list, current_month_list_values):
        if month == current_month_list[0]:
            start_row = 2
            a = current_month_list_values[0] + 2
            finish_row = a
        elif month == current_month_list[1]:
            a = current_month_list_values[0] + 2
            b = current_month_list_values[1] + a
            start_row = a
            finish_row = b
        elif month == current_month_list[2]:
            a = current_month_list_values[0] + 2
            b = current_month_list_values[1] + a
            c = current_month_list_values[2] + b
            start_row = b
            finish_row = c
        elif month == current_month_list[3]:
            a = current_month_list_values[0] + 2
            b = current_month_list_values[1] + a
            c = current_month_list_values[2] + b
            d = current_month_list_values[3] + c
            start_row = c
            finish_row =d
        elif month == current_month_list[4]:
            a = current_month_list_values[0] + 2
            b = current_month_list_values[1] + a
            c = current_month_list_values[2] + b
            d = current_month_list_values[3] + c
            e = current_month_list_values[4] + d
            start_row = d
            finish_row = e
        elif month == current_month_list[5]:
            a = current_month_list_values[0] + 2
            b = current_month_list_values[1] + a
            c = current_month_list_values[2] + b
            d = current_month_list_values[3] + c
            e = current_month_list_values[4] + d
            f = current_month_list_values[5] + e
            start_row = e
            finish_row = f
        elif month == current_month_list[6]:
            a = current_month_list_values[0] + 2
            b = current_month_list_values[1] + a
            c = current_month_list_values[2] + b
            d = current_month_list_values[3] + c
            e = current_month_list_values[4] + d
            f = current_month_list_values[5] + e
            g = current_month_list_values[6] + f
            start_row = f
            finish_row = g
        elif month == current_month_list[7]:
            a = current_month_list_values[0] + 2
            b = current_month_list_values[1] + a
            c = current_month_list_values[2] + b
            d = current_month_list_values[3] + c
            e = current_month_list_values[4] + d
            f = current_month_list_values[5] + e
            g = current_month_list_values[6] + f
            h = current_month_list_values[7] + g
            start_row = g
            finish_row = h
        elif month == current_month_list[8]:
            a = current_month_list_values[0] + 2
            b = current_month_list_values[1] + a
            c = current_month_list_values[2] + b
            d = current_month_list_values[3] + c
            e = current_month_list_values[4] + d
            f = current_month_list_values[5] + e
            g = current_month_list_values[6] + f
            h = current_month_list_values[7] + g
            i = current_month_list_values[8] + h
            start_row = h
            finish_row = i
        elif month == current_month_list[9]:
            a = current_month_list_values[0] + 2
            b = current_month_list_values[1] + a
            c = current_month_list_values[2] + b
            d = current_month_list_values[3] + c
            e = current_month_list_values[4] + d
            f = current_month_list_values[5] + e
            g = current_month_list_values[6] + f
            h = current_month_list_values[7] + g
            i = current_month_list_values[8] + h
            j = current_month_list_values[9] + i
            start_row = i
            finish_row = j
        elif month == current_month_list[10]:
            a = current_month_list_values[0] + 2
            b = current_month_list_values[1] + a
            c = current_month_list_values[2] + b
            d = current_month_list_values[3] + c
            e = current_month_list_values[4] + d
            f = current_month_list_values[5] + e
            g = current_month_list_values[6] + f
            h = current_month_list_values[7] + g
            i = current_month_list_values[8] + h
            j = current_month_list_values[9] + i
            k = current_month_list_values[10] + j
            start_row = j
            finish_row = k
        elif month == current_month_list[11]:
            a = current_month_list_values[0] + 2
            b = current_month_list_values[1] + a
            c = current_month_list_values[2] + b
            d = current_month_list_values[3] + c
            e = current_month_list_values[4] + d
            f = current_month_list_values[5] + e
            g = current_month_list_values[6] + f
            h = current_month_list_values[7] + g
            i = current_month_list_values[8] + h
            j = current_month_list_values[9] + i
            k = current_month_list_values[10] + j
            l = current_month_list_values[s11] + k
        return [start_row, finish_row]


    def get_students_scheduled_for_classes(work_book, teacher, row_number):
        sheet = work_book[teacher]
        student_list = []
        for value in range(1, 16):
            if sheet.cell(row=row_number, column=value).value == None:
                student_list.append(' ')
            else:
                student_list.append(sheet.cell(row=row_number, column=value).value)
        for i in range(17, 18):
            if sheet.cell(row=row_number, column=value).value == None:
                student_list.append(' ')
            else:
                student_list.append(sheet.cell(row=row_number, column=value).value)

        return student_list

# ====================================================================================================================================================
class GUI_Input_Output:
    
    def __init__(self, master):

        master.title('TLI Search for Available Teachers by Date and Time')
        # master.resizable(False, False)
        master.configure(background = 'black', padx = 2, pady = 2)

        self.style = ttk.Style()
        self.style.configure('TFrame', background = 'black')
        self.style.configure('TButton', background = 'dark blue', foreground = 'gold')
        self.style.configure('TLabel', background = 'black', foreground = 'gold', font = ('Arial', 11))
        # Note: the header label inherits the overall label's color
        self.style.configure('Header.TLabel', font = ('Arial', 18, 'bold'))
# ====================================================================================================================================================
# tab 1: GUI search for available teachers ***********************************************************************************************************
        self.tabControl = ttk.Notebook(master)
        tab1 = ttk.Frame(self.tabControl)
        self.tabControl.add(tab1, text='Search Available Teachers')
        self.tabControl.pack(expand=1, fill='both')

        frame_header = ttk.Frame(tab1)
        frame_header.pack()

        ttk.Label(frame_header, text = 'Search for Available Teachers', style = 'Header.TLabel').grid(row = 0, column = 1, pady = 15)
        ttk.Label(frame_header, wraplength = 750,
                  text = ("This program will search for available teachers by date and time."
                          "You can search for available teachers in up to four time slots."
                          "If you would like to search for available teachers in fewer time slots,"
                          "leave the remaining variables blank.")).grid(row = 1, column = 1, padx = 20)


        self.frame_content1 = ttk.Frame(tab1)
        self.frame_content1.pack()        
        self.logo2 = PhotoImage(file = 'TLI_logo_color_resize.png')
        ttk.Label(self.frame_content1, image=self.logo2).grid(row=0, column=0, rowspan=9, padx = 20, pady = 2)
        ttk.Label(self.frame_content1, image=self.logo2).grid(row=0, column=3, rowspan=9, padx = 20, pady = 2)
        
        # Entry comboboxes for user to enter dates and times
        ttk.Label(self.frame_content1, text = "Date 1").grid(row = 0, column = 1, padx = 5, pady = 5)
        date1 = StringVar()
        self.entry_combobox1 = ttk.Combobox(self.frame_content1, textvariable = date1)
        self.entry_combobox1.grid(row=1, column=1, padx = 5, pady = 5)
        self.entry_combobox1.config(values = Get_Data_From_Excel.get_date_list())
        createToolTip(self.entry_combobox1, 'First Date.')

        ttk.Label(self.frame_content1, text = "Time1").grid(row = 2, column = 1, padx = 5, pady = 5)
        time1 = StringVar()
        self.entry_combobox2 = ttk.Combobox(self.frame_content1, textvariable = time1)
        self.entry_combobox2.grid(row=3, column=1, padx = 5, pady = 5)
        self.entry_combobox2.config(values = Get_Data_From_Excel.get_time_list())
        createToolTip(self.entry_combobox2, 'First Time.')

        ttk.Label(self.frame_content1, text = "Date 2").grid(row = 4, column = 1, padx = 5, pady = 5)
        date2= StringVar()
        self.entry_combobox3 = ttk.Combobox(self.frame_content1, textvariable = date2)
        self.entry_combobox3.grid(row=5, column=1, padx = 5, pady = 5)
        self.entry_combobox3.config(values = Get_Data_From_Excel.get_date_list())
        createToolTip(self.entry_combobox3, 'Second Date.')

        ttk.Label(self.frame_content1, text = "Time2").grid(row = 6, column = 1, padx = 5, pady = 5)
        time2 = StringVar()
        self.entry_combobox4 = ttk.Combobox(self.frame_content1, textvariable = time2)
        self.entry_combobox4.grid(row=7, column=1, padx = 5, pady = 5)
        self.entry_combobox4.config(values = Get_Data_From_Excel.get_time_list())
        createToolTip(self.entry_combobox4, 'Second Time.')

        ttk.Label(self.frame_content1, text = "Date 3").grid(row = 0, column = 2, padx = 5, pady = 5)
        date3 = StringVar()
        self.entry_combobox5 = ttk.Combobox(self.frame_content1, textvariable = date3)
        self.entry_combobox5.grid(row=1, column=2, padx = 5, pady = 5)
        self.entry_combobox5.config(values = Get_Data_From_Excel.get_date_list())
        createToolTip(self.entry_combobox5, 'Third Date.')

        ttk.Label(self.frame_content1, text = "Time3").grid(row = 2, column = 2, padx = 5, pady = 5)
        time3 = StringVar()
        self.entry_combobox6 = ttk.Combobox(self.frame_content1, textvariable = time3)
        self.entry_combobox6.grid(row=3, column=2, padx = 5, pady = 5)
        self.entry_combobox6.config(values = Get_Data_From_Excel.get_time_list())
        createToolTip(self.entry_combobox6, 'Third Time.')

        ttk.Label(self.frame_content1, text = "Date 4").grid(row = 4, column = 2, padx = 5, pady = 5)
        date4 = StringVar()
        self.entry_combobox7 = ttk.Combobox(self.frame_content1, textvariable = date4)
        self.entry_combobox7.grid(row=5, column=2, padx = 5, pady = 5)
        self.entry_combobox7.config(values = Get_Data_From_Excel.get_date_list())
        createToolTip(self.entry_combobox7, 'Fourth Date.')

        ttk.Label(self.frame_content1, text = "Time4").grid(row = 6, column = 2, padx = 5, pady = 5)
        time4 = StringVar()
        self.entry_combobox8 = ttk.Combobox(self.frame_content1, textvariable = time4)
        self.entry_combobox8.grid(row=7, column=2, padx = 5, pady = 5)
        self.entry_combobox8.config(values = Get_Data_From_Excel.get_time_list())
        createToolTip(self.entry_combobox8, 'Fourth Time.')

        ttk.Button(self.frame_content1, text = 'Search', command = self.search_free_teachers).grid(row = 8, column = 1, padx = 5, pady = 10)
        ttk.Button(self.frame_content1, text = 'Clear', command = self.clear).grid(row = 8, column = 2, padx = 5, pady = 10)

        ttk.Label(self.frame_content1, text = "Teachers available:").grid(row = 9, column = 1, padx = 5, pady = 5)
        ttk.Label(self.frame_content1, text = "Contact Details:").grid(row = 9, column = 2, padx = 5, pady = 5)

        self.text_names = Text(self.frame_content1, width = 40, height = 15)
        self.text_names.grid(row=10, column=1, sticky = 'nsew', padx = 5, pady = 5)
        createToolTip(self.text_names, 'Names of Available Teachers.')

        self.text_emails = Text(self.frame_content1, width = 40, height = 15)
        self.text_emails.grid(row=10, column=2, sticky = 'nsew', padx = 5, pady = 5)
        createToolTip(self.text_emails, 'Email Contacts for Available Teachers.')
# Tab 2: GUI email teachers scheduling information ==================================================================================================================
        tab2 = ttk.Frame(self.tabControl)
        self.tabControl.add(tab2, text='Automated Scheduling Emails')


        frame_header2 = ttk.Frame(tab2)
        frame_header2.pack()

        ttk.Label(frame_header2, text = 'Scheduling Reminder', style = 'Header.TLabel').grid(row = 0, column = 1, padx = 25, pady = 15)
        ttk.Label(frame_header2, wraplength = 750,
                  text = ("This will send out scheduling reminders to TLI teachers for tomorrow's classes.  "
                          "If you would like to leave an additional message for the teachers, enter the text here."),
                  style = 'Header.TLabel').grid(row = 1, column = 1, padx = 30, pady = 30)


        self.frame_content2 = ttk.Frame(tab2)
        self.frame_content2.pack()
        
        self.logo = PhotoImage(file = 'TLI_logo_color_resize.png')
        ttk.Label(self.frame_content2, image=self.logo).grid(row=0, column=0, rowspan=4, padx = 20, pady = 2)
        ttk.Label(self.frame_content2, image=self.logo).grid(row=0, column=3, rowspan=4, padx = 20, pady = 2)
        
        ttk.Label(self.frame_content2, text = 'Your email: ').grid(row = 0, column = 1, padx = 20, sticky = 'sw')
        ttk.Label(self.frame_content2, text = 'Your password: ').grid(row = 0, column = 2, padx = 20, sticky = 'sw')
        ttk.Label(self.frame_content2, text = 'Additional Message: ').grid(row = 2, column = 1, padx = 20, sticky = 'sw')

        self.entry_email = ttk.Entry(self.frame_content2, width = 24, font = ('Arial', 10))
        self.entry_password = ttk.Entry(self.frame_content2, width = 24, font = ('Arial', 10), show = '*')
        self.text_message = Text(self.frame_content2, width = 50, height = 10, pady = 25, font = ('Arial', 10))

        self.entry_email.grid(row = 1, column = 1, padx = 20)
        createToolTip(self.entry_email, "Enter your email address")
        self.entry_password.grid(row = 1, column = 2, padx = 20)
        createToolTip(self.entry_password, "Enter your password")
        self.text_message.grid(row = 3, column = 1, columnspan = 2, padx = 30, sticky = 'sw')
        createToolTip(self.text_message, "You can enter an additional message (optional)")

        ttk.Button(self.frame_content2, text = 'Submit', command = self.submit_schedules_for_email).grid(row = 4, column = 1, padx = 20, pady = 30, sticky = 'e')
        ttk.Button(self.frame_content2, text = 'Clear', command = self.clear2).grid(row = 4, column = 2, padx = 20, pady = 30, sticky = 'w')

# GUI for looking up monthly schedules ===========================================================================================================================================
        tab3 = ttk.Frame(self.tabControl)
        self.tabControl.add(tab3, text="TLI Search for Teacher's Monthly Schedule")
        self.tabControl.pack(expand=1, fill='both')
        frame_header3 = ttk.Frame(tab3)
        frame_header3.pack()

        entryW = 10
        entryW2 = 11
        entryW3 = 12

        sheetList = Get_Data_From_Excel.get_teacher_list()
        current_month_list_and_values = Get_Data_From_Excel.get_current_month_list_and_month_list_values()
        current_month_list = current_month_list_and_values[0]
        current_month_list_values = current_month_list_and_values[1]

        teacher_Name_List =  sheetList # redundancy?

        #ttk.Label(frame_header3, text = "Search for Teacher's Monthly Schedule").grid(row = 0, column = 1)

        self.frame_content = ttk.Frame(tab3)
        self.frame_content.pack()        
        

        ttk.Label(self.frame_content, text = "Teacher's Name").grid(row = 0, column = 4, columnspan = 3, padx = 5, pady = 1)
        date1 = StringVar()
        self.entry_combobox_1 = ttk.Combobox(self.frame_content, textvariable = date1)
        self.entry_combobox_1.grid(row=1, column=4, columnspan = 3, padx = 5, pady = 1)
        self.entry_combobox_1.config(values = sheetList)

        ttk.Label(self.frame_content, text = "Month").grid(row = 0, column = 9, columnspan = 3, padx = 5, pady = 1)
        time1 = StringVar()
        self.entry_combobox_2 = ttk.Combobox(self.frame_content, textvariable = time1)
        self.entry_combobox_2.grid(row=1, column=9, columnspan = 3, padx = 5, pady = 1)
        self.entry_combobox_2.config(values = current_month_list)

        ttk.Button(self.frame_content, text = 'Search', command = self.search_monthly_schedule).grid(row = 2, column = 4, columnspan = 3, padx = 5, pady = 1)
        ttk.Button(self.frame_content, text = 'Clear', command = self.clear3).grid(row = 2, column = 9, columnspan = 3, padx = 5, pady = 1)


        self.entry_cell_one = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
        
        entry_cell_list = [self.entry_cell_one, self.entry_cell_two, self.entry_cell_three, self.entry_cell_four, self.entry_cell_five,
                           self.entry_cell_six, self.entry_cell_seven, self.entry_cell_eight, self.entry_cell_nine, self.entry_cell_ten,
                           self.entry_cell_eleven, self.entry_cell_twelve, self.entry_cell_thirteen, self.entry_cell_fourteen,
                           self.entry_cell_fifteen, self.entry_cell_sixteen]

        self.entry_cell_one_2 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_2 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_2 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_2 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_2 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_2 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_2 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_2 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_2 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_2 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_2 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_2 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_2 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_2 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_2 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_2 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
        
        entry_cell_list_2 = [self.entry_cell_one_2, self.entry_cell_two_2, self.entry_cell_three_2, self.entry_cell_four_2, self.entry_cell_five_2,
                             self.entry_cell_six_2, self.entry_cell_seven_2, self.entry_cell_eight_2, self.entry_cell_nine_2, self.entry_cell_ten_2,
                             self.entry_cell_eleven_2, self.entry_cell_twelve_2, self.entry_cell_thirteen_2, self.entry_cell_fourteen_2,
                             self.entry_cell_fifteen_2, self.entry_cell_sixteen_2]

        self.entry_cell_one_3 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_3 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_3 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_3 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_3 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_3 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_3 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_3 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_3 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_3 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_3 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_3 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_3 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_3 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_3 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_3 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
        
        entry_cell_list_3 = [self.entry_cell_one_3, self.entry_cell_two_3, self.entry_cell_three_3, self.entry_cell_four_3, self.entry_cell_five_3,
                             self.entry_cell_six_3, self.entry_cell_seven_3, self.entry_cell_eight_3, self.entry_cell_nine_3, self.entry_cell_ten_3,
                             self.entry_cell_eleven_3, self.entry_cell_twelve_3, self.entry_cell_thirteen_3, self.entry_cell_fourteen_3,
                             self.entry_cell_fifteen_3, self.entry_cell_sixteen_3]

        self.entry_cell_one_4 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_4 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_4 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_4 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_4 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_4 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_4 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_4 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_4 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_4 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_4 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_4 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_4 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_4 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_4 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_4 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
                
        entry_cell_list_4 = [self.entry_cell_one_4, self.entry_cell_two_4, self.entry_cell_three_4, self.entry_cell_four_4, self.entry_cell_five_4,
                                  self.entry_cell_six_4, self.entry_cell_seven_4, self.entry_cell_eight_4, self.entry_cell_nine_4, self.entry_cell_ten_4,
                                  self.entry_cell_eleven_4, self.entry_cell_twelve_4, self.entry_cell_thirteen_4, self.entry_cell_fourteen_4,
                                  self.entry_cell_fifteen_4, self.entry_cell_sixteen_4]


        self.entry_cell_one_5 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_5 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_5 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_5 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_5 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_5 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_5 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_5 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_5 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_5 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_5 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_5 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_5 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_5 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_5 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_5 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
                
        entry_cell_list_5 = [self.entry_cell_one_5, self.entry_cell_two_5, self.entry_cell_three_5, self.entry_cell_four_5, self.entry_cell_five_5,
                                  self.entry_cell_six_5, self.entry_cell_seven_5, self.entry_cell_eight_5, self.entry_cell_nine_5, self.entry_cell_ten_5,
                                  self.entry_cell_eleven_5, self.entry_cell_twelve_5, self.entry_cell_thirteen_5, self.entry_cell_fourteen_5,
                                  self.entry_cell_fifteen_5, self.entry_cell_sixteen_5]


        self.entry_cell_one_6 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_6 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_6 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_6 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_6 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_6 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_6 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_6 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_6 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_6 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_6 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_6 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_6 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_6 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_6 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_6 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
                
        entry_cell_list_6 = [self.entry_cell_one_6, self.entry_cell_two_6, self.entry_cell_three_6, self.entry_cell_four_6, self.entry_cell_five_6,
                                  self.entry_cell_six_6, self.entry_cell_seven_6, self.entry_cell_eight_6, self.entry_cell_nine_6, self.entry_cell_ten_6,
                                  self.entry_cell_eleven_6, self.entry_cell_twelve_6, self.entry_cell_thirteen_6, self.entry_cell_fourteen_6,
                                  self.entry_cell_fifteen_6, self.entry_cell_sixteen_6]

        self.entry_cell_one_7 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_7 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_7 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_7 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_7 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_7 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_7 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_7 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_7 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_7 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_7 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_7 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_7 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_7 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_7 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_7 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
                
        entry_cell_list_7 = [self.entry_cell_one_7, self.entry_cell_two_7, self.entry_cell_three_7, self.entry_cell_four_7, self.entry_cell_five_7,
                                  self.entry_cell_six_7, self.entry_cell_seven_7, self.entry_cell_eight_7, self.entry_cell_nine_7, self.entry_cell_ten_7,
                                  self.entry_cell_eleven_7, self.entry_cell_twelve_7, self.entry_cell_thirteen_7, self.entry_cell_fourteen_7,
                                  self.entry_cell_fifteen_7, self.entry_cell_sixteen_7]

        self.entry_cell_one_8 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_8 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_8 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_8 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_8 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_8 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_8 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_8 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_8 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_8 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_8 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_8 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_8 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_8 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_8 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_8 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
                
        entry_cell_list_8 = [self.entry_cell_one_8, self.entry_cell_two_8, self.entry_cell_three_8, self.entry_cell_four_8, self.entry_cell_five_8,
                                  self.entry_cell_six_8, self.entry_cell_seven_8, self.entry_cell_eight_8, self.entry_cell_nine_8, self.entry_cell_ten_8,
                                  self.entry_cell_eleven_8, self.entry_cell_twelve_8, self.entry_cell_thirteen_8, self.entry_cell_fourteen_8,
                                  self.entry_cell_fifteen_8, self.entry_cell_sixteen_8]

        self.entry_cell_one_9 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_9 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_9 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_9 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_9 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_9 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_9 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_9 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_9 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_9 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_9 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_9 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_9 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_9 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_9 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_9 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
                
        entry_cell_list_9 = [self.entry_cell_one_9, self.entry_cell_two_9, self.entry_cell_three_9, self.entry_cell_four_9, self.entry_cell_five_9,
                                  self.entry_cell_six_9, self.entry_cell_seven_9, self.entry_cell_eight_9, self.entry_cell_nine_9, self.entry_cell_ten_9,
                                  self.entry_cell_eleven_9, self.entry_cell_twelve_9, self.entry_cell_thirteen_9, self.entry_cell_fourteen_9,
                                  self.entry_cell_fifteen_9, self.entry_cell_sixteen_9]


        self.entry_cell_one_10 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_10 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_10 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_10 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_10 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_10 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_10 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_10 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_10 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_10 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_10 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_10 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_10 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_10 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_10 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_10 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
                
        entry_cell_list_10 = [self.entry_cell_one_10, self.entry_cell_two_10, self.entry_cell_three_10, self.entry_cell_four_10, self.entry_cell_five_10,
                                  self.entry_cell_six_10, self.entry_cell_seven_10, self.entry_cell_eight_10, self.entry_cell_nine_10, self.entry_cell_ten_10,
                                  self.entry_cell_eleven_10, self.entry_cell_twelve_10, self.entry_cell_thirteen_10, self.entry_cell_fourteen_10,
                                  self.entry_cell_fifteen_10, self.entry_cell_sixteen_10]

        self.entry_cell_one_11 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_11 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_11 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_11 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_11 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_11 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_11 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_11 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_11 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_11 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_11 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_11 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_11 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_11 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_11 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_11 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
                
        entry_cell_list_11 = [self.entry_cell_one_11, self.entry_cell_two_11, self.entry_cell_three_11, self.entry_cell_four_11, self.entry_cell_five_11,
                                  self.entry_cell_six_11, self.entry_cell_seven_11, self.entry_cell_eight_11, self.entry_cell_nine_11, self.entry_cell_ten_11,
                                  self.entry_cell_eleven_11, self.entry_cell_twelve_11, self.entry_cell_thirteen_11, self.entry_cell_fourteen_11,
                                  self.entry_cell_fifteen_11, self.entry_cell_sixteen_11]

        self.entry_cell_one_12 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_12 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_12 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_12 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_12 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_12 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_12 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_12 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_12 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_12 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_12 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_12 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_12 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_12 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_12 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_12 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
                
        entry_cell_list_12 = [self.entry_cell_one_12, self.entry_cell_two_12, self.entry_cell_three_12, self.entry_cell_four_12, self.entry_cell_five_12,
                                  self.entry_cell_six_12, self.entry_cell_seven_12, self.entry_cell_eight_12, self.entry_cell_nine_12, self.entry_cell_ten_12,
                                  self.entry_cell_eleven_12, self.entry_cell_twelve_12, self.entry_cell_thirteen_12, self.entry_cell_fourteen_12,
                                  self.entry_cell_fifteen_12, self.entry_cell_sixteen_12]

        self.entry_cell_one_13 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_13 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_13 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_13 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_13 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_13 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_13 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_13 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_13 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_13 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_13 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_13 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_13 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_13 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_13 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_13 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
                
        entry_cell_list_13 = [self.entry_cell_one_13, self.entry_cell_two_13, self.entry_cell_three_13, self.entry_cell_four_13, self.entry_cell_five_13,
                                  self.entry_cell_six_13, self.entry_cell_seven_13, self.entry_cell_eight_13, self.entry_cell_nine_13, self.entry_cell_ten_13,
                                  self.entry_cell_eleven_13, self.entry_cell_twelve_13, self.entry_cell_thirteen_13, self.entry_cell_fourteen_13,
                                  self.entry_cell_fifteen_13, self.entry_cell_sixteen_13]

        self.entry_cell_one_14 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_14 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_14 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_14 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_14 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_14 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_14 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_14 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_14 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_14 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_14 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_14 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_14 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_14 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_14 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_14 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
                
        entry_cell_list_14 = [self.entry_cell_one_14, self.entry_cell_two_14, self.entry_cell_three_14, self.entry_cell_four_14, self.entry_cell_five_14,
                                  self.entry_cell_six_14, self.entry_cell_seven_14, self.entry_cell_eight_14, self.entry_cell_nine_14, self.entry_cell_ten_14,
                                  self.entry_cell_eleven_14, self.entry_cell_twelve_14, self.entry_cell_thirteen_14, self.entry_cell_fourteen_14,
                                  self.entry_cell_fifteen_14, self.entry_cell_sixteen_14]


        self.entry_cell_one_15 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_15 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_15 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_15 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_15 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_15 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_15 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_15 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_15 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_15 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_15 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_15 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_15 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_15 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_15 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_15 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
                
        entry_cell_list_15 = [self.entry_cell_one_15, self.entry_cell_two_15, self.entry_cell_three_15, self.entry_cell_four_15, self.entry_cell_five_15,
                                  self.entry_cell_six_15, self.entry_cell_seven_15, self.entry_cell_eight_15, self.entry_cell_nine_15, self.entry_cell_ten_15,
                                  self.entry_cell_eleven_15, self.entry_cell_twelve_15, self.entry_cell_thirteen_15, self.entry_cell_fourteen_15,
                                  self.entry_cell_fifteen_15, self.entry_cell_sixteen_15]

        self.entry_cell_one_16 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_16 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_16 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_16 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_16 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_16 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_16 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_16 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_16 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_16 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_16 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_16 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_16 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_16 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_16 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_16 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
                
        entry_cell_list_16 = [self.entry_cell_one_16, self.entry_cell_two_16, self.entry_cell_three_16, self.entry_cell_four_16, self.entry_cell_five_16,
                                  self.entry_cell_six_16, self.entry_cell_seven_16, self.entry_cell_eight_16, self.entry_cell_nine_16, self.entry_cell_ten_16,
                                  self.entry_cell_eleven_16, self.entry_cell_twelve_16, self.entry_cell_thirteen_16, self.entry_cell_fourteen_16,
                                  self.entry_cell_fifteen_16, self.entry_cell_sixteen_16]

        self.entry_cell_one_17 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_17 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_17 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_17 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_17 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_17 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_17 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_17 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_17 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_17 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_17 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_17 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_17 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_17 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_17 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_17 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
                
        entry_cell_list_17 = [self.entry_cell_one_17, self.entry_cell_two_17, self.entry_cell_three_17, self.entry_cell_four_17, self.entry_cell_five_17,
                                  self.entry_cell_six_17, self.entry_cell_seven_17, self.entry_cell_eight_17, self.entry_cell_nine_17, self.entry_cell_ten_17,
                                  self.entry_cell_eleven_17, self.entry_cell_twelve_17, self.entry_cell_thirteen_17, self.entry_cell_fourteen_17,
                                  self.entry_cell_fifteen_17, self.entry_cell_sixteen_17]

        self.entry_cell_one_18 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_18 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_18 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_18 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_18 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_18 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_18 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_18 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_18 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_18 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_18 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_18 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_18 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_18 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_18 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_18 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
                
        entry_cell_list_18 = [self.entry_cell_one_18, self.entry_cell_two_18, self.entry_cell_three_18, self.entry_cell_four_18, self.entry_cell_five_18,
                                  self.entry_cell_six_18, self.entry_cell_seven_18, self.entry_cell_eight_18, self.entry_cell_nine_18, self.entry_cell_ten_18,
                                  self.entry_cell_eleven_18, self.entry_cell_twelve_18, self.entry_cell_thirteen_18, self.entry_cell_fourteen_18,
                                  self.entry_cell_fifteen_18, self.entry_cell_sixteen_18]

        self.entry_cell_one_19 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_19 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_19 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_19 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_19 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_19 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_19 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_19 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_19 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_19 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_19 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_19 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_19 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_19 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_19 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_19 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
                
        entry_cell_list_19 = [self.entry_cell_one_19, self.entry_cell_two_19, self.entry_cell_three_19, self.entry_cell_four_19, self.entry_cell_five_19,
                                  self.entry_cell_six_19, self.entry_cell_seven_19, self.entry_cell_eight_19, self.entry_cell_nine_19, self.entry_cell_ten_19,
                                  self.entry_cell_eleven_19, self.entry_cell_twelve_19, self.entry_cell_thirteen_19, self.entry_cell_fourteen_19,
                                  self.entry_cell_fifteen_19, self.entry_cell_sixteen_19]

        self.entry_cell_one_20 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_20 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_20 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_20 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_20 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_20 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_20 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_20 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_20 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_20 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_20 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_20 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_20 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_20 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_20 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_20 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
                
        entry_cell_list_20 = [self.entry_cell_one_20, self.entry_cell_two_20, self.entry_cell_three_20, self.entry_cell_four_20, self.entry_cell_five_20,
                                  self.entry_cell_six_20, self.entry_cell_seven_20, self.entry_cell_eight_20, self.entry_cell_nine_20, self.entry_cell_ten_20,
                                  self.entry_cell_eleven_20, self.entry_cell_twelve_20, self.entry_cell_thirteen_20, self.entry_cell_fourteen_20,
                                  self.entry_cell_fifteen_20, self.entry_cell_sixteen_20]

        self.entry_cell_one_21 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_21 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_21 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_21 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_21 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_21 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_21 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_21 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_21 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_21 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_21 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_21 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_21 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_21 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_21 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_21 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
                
        entry_cell_list_21 = [self.entry_cell_one_21, self.entry_cell_two_21, self.entry_cell_three_21, self.entry_cell_four_21, self.entry_cell_five_21,
                                  self.entry_cell_six_21, self.entry_cell_seven_21, self.entry_cell_eight_21, self.entry_cell_nine_21, self.entry_cell_ten_21,
                                  self.entry_cell_eleven_21, self.entry_cell_twelve_21, self.entry_cell_thirteen_21, self.entry_cell_fourteen_21,
                                  self.entry_cell_fifteen_21, self.entry_cell_sixteen_21]

        self.entry_cell_one_22 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_22 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_22 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_22 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_22 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_22 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_22 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_22 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_22 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_22 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_22 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_22 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_22 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_22 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_22 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_22 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
                
        entry_cell_list_22 = [self.entry_cell_one_22, self.entry_cell_two_22, self.entry_cell_three_22, self.entry_cell_four_22, self.entry_cell_five_22,
                                  self.entry_cell_six_22, self.entry_cell_seven_22, self.entry_cell_eight_22, self.entry_cell_nine_22, self.entry_cell_ten_22,
                                  self.entry_cell_eleven_22, self.entry_cell_twelve_22, self.entry_cell_thirteen_22, self.entry_cell_fourteen_22,
                                  self.entry_cell_fifteen_22, self.entry_cell_sixteen_22]

        self.entry_cell_one_23 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_23 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_23 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_23 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_23 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_23 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_23 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_23 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_23 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_23 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_23 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_23 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_23 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_23 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_23 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_23 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
                
        entry_cell_list_23 = [self.entry_cell_one_23, self.entry_cell_two_23, self.entry_cell_three_23, self.entry_cell_four_23, self.entry_cell_five_23,
                                  self.entry_cell_six_23, self.entry_cell_seven_23, self.entry_cell_eight_23, self.entry_cell_nine_23, self.entry_cell_ten_23,
                                  self.entry_cell_eleven_23, self.entry_cell_twelve_23, self.entry_cell_thirteen_23, self.entry_cell_fourteen_23,
                                  self.entry_cell_fifteen_23, self.entry_cell_sixteen_23]

        self.entry_cell_one_24 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_24 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_24 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_24 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_24 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_24 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_24 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_24 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_24 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_24 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_24 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_24 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_24 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_24 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_24 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_24 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
                
        entry_cell_list_24 = [self.entry_cell_one_24, self.entry_cell_two_24, self.entry_cell_three_24, self.entry_cell_four_24, self.entry_cell_five_24,
                                  self.entry_cell_six_24, self.entry_cell_seven_24, self.entry_cell_eight_24, self.entry_cell_nine_24, self.entry_cell_ten_24,
                                  self.entry_cell_eleven_24, self.entry_cell_twelve_24, self.entry_cell_thirteen_24, self.entry_cell_fourteen_24,
                                  self.entry_cell_fifteen_24, self.entry_cell_sixteen_24]

        self.entry_cell_one_25 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_25 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_25 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_25 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_25 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_25 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_25 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_25 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_25 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_25 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_25 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_25 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_25 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_25 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_25 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_25 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
                
        entry_cell_list_25 = [self.entry_cell_one_25, self.entry_cell_two_25, self.entry_cell_three_25, self.entry_cell_four_25, self.entry_cell_five_25,
                                  self.entry_cell_six_25, self.entry_cell_seven_25, self.entry_cell_eight_25, self.entry_cell_nine_25, self.entry_cell_ten_25,
                                  self.entry_cell_eleven_25, self.entry_cell_twelve_25, self.entry_cell_thirteen_25, self.entry_cell_fourteen_25,
                                  self.entry_cell_fifteen_25, self.entry_cell_sixteen_25]

        self.entry_cell_one_26 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_26 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_26 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_26 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_26 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_26 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_26 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_26 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_26 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_26 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_26 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_26 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_26 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_26 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_26 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_26 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
                
        entry_cell_list_26 = [self.entry_cell_one_26, self.entry_cell_two_26, self.entry_cell_three_26, self.entry_cell_four_26, self.entry_cell_five_26,
                                  self.entry_cell_six_26, self.entry_cell_seven_26, self.entry_cell_eight_26, self.entry_cell_nine_26, self.entry_cell_ten_26,
                                  self.entry_cell_eleven_26, self.entry_cell_twelve_26, self.entry_cell_thirteen_26, self.entry_cell_fourteen_26,
                                  self.entry_cell_fifteen_26, self.entry_cell_sixteen_26]

        self.entry_cell_one_27 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_27 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_27 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_27 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_27 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_27 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_27 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_27 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_27 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_27 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_27 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_27 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_27 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_27 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_27 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_27 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
                
        entry_cell_list_27 = [self.entry_cell_one_27, self.entry_cell_two_27, self.entry_cell_three_27, self.entry_cell_four_27, self.entry_cell_five_27,
                                  self.entry_cell_six_27, self.entry_cell_seven_27, self.entry_cell_eight_27, self.entry_cell_nine_27, self.entry_cell_ten_27,
                                  self.entry_cell_eleven_27, self.entry_cell_twelve_27, self.entry_cell_thirteen_27, self.entry_cell_fourteen_27,
                                  self.entry_cell_fifteen_27, self.entry_cell_sixteen_27]

        self.entry_cell_one_28 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_28 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_28 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_28 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_28 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_28 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_28 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_28 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_28 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_28 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_28 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_28 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_28 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_28 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_28 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_28 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
                
        entry_cell_list_28 = [self.entry_cell_one_28, self.entry_cell_two_28, self.entry_cell_three_28, self.entry_cell_four_28, self.entry_cell_five_28,
                                  self.entry_cell_six_28, self.entry_cell_seven_28, self.entry_cell_eight_28, self.entry_cell_nine_28, self.entry_cell_ten_28,
                                  self.entry_cell_eleven_28, self.entry_cell_twelve_28, self.entry_cell_thirteen_28, self.entry_cell_fourteen_28,
                                  self.entry_cell_fifteen_28, self.entry_cell_sixteen_28]

        self.entry_cell_one_29 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_29 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_29 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_29 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_29 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_29 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_29 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_29 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_29 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_29 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_29 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_29 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_29 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_29 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_29 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_29 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
                
        entry_cell_list_29 = [self.entry_cell_one_29, self.entry_cell_two_29, self.entry_cell_three_29, self.entry_cell_four_29, self.entry_cell_five_29,
                                  self.entry_cell_six_29, self.entry_cell_seven_29, self.entry_cell_eight_29, self.entry_cell_nine_29, self.entry_cell_ten_29,
                                  self.entry_cell_eleven_29, self.entry_cell_twelve_29, self.entry_cell_thirteen_29, self.entry_cell_fourteen_29,
                                  self.entry_cell_fifteen_29, self.entry_cell_sixteen_29]

        self.entry_cell_one_30 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_30 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_30 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_30 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_30 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_30 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_30 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_30 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_30 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_30 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_30 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_30 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_30 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_30 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_30 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_30 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
                
        entry_cell_list_30 = [self.entry_cell_one_30, self.entry_cell_two_30, self.entry_cell_three_30, self.entry_cell_four_30, self.entry_cell_five_30,
                                  self.entry_cell_six_30, self.entry_cell_seven_30, self.entry_cell_eight_30, self.entry_cell_nine_30, self.entry_cell_ten_30,
                                  self.entry_cell_eleven_30, self.entry_cell_twelve_30, self.entry_cell_thirteen_30, self.entry_cell_fourteen_30,
                                  self.entry_cell_fifteen_30, self.entry_cell_sixteen_30]


        self.entry_cell_one_31 = ttk.Entry(self.frame_content, width = entryW2, font = ('Arial', 10))
        self.entry_cell_two_31 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_three_31 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_four_31 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_five_31 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_six_31 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_seven_31 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eight_31 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_nine_31 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_ten_31 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_eleven_31 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_twelve_31 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_thirteen_31 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fourteen_31 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_fifteen_31 = ttk.Entry(self.frame_content, width = entryW, font = ('Arial', 10))
        self.entry_cell_sixteen_31 = ttk.Entry(self.frame_content, width = entryW3, font = ('Arial', 10))
                
        entry_cell_list_31 = [self.entry_cell_one_31, self.entry_cell_two_31, self.entry_cell_three_31, self.entry_cell_four_31, self.entry_cell_five_31,
                                  self.entry_cell_six_31, self.entry_cell_seven_31, self.entry_cell_eight_31, self.entry_cell_nine_31, self.entry_cell_ten_31,
                                  self.entry_cell_eleven_31, self.entry_cell_twelve_31, self.entry_cell_thirteen_31, self.entry_cell_fourteen_31,
                                  self.entry_cell_fifteen_31, self.entry_cell_sixteen_31]


        entry_cell_list_list = [entry_cell_list, entry_cell_list_2, entry_cell_list_3, entry_cell_list_4, entry_cell_list_5, entry_cell_list_6,
                                entry_cell_list_7, entry_cell_list_8, entry_cell_list_9, entry_cell_list_10, entry_cell_list_11, entry_cell_list_11,
                                entry_cell_list_12, entry_cell_list_13, entry_cell_list_14, entry_cell_list_15, entry_cell_list_16, entry_cell_list_17,
                                entry_cell_list_18, entry_cell_list_19, entry_cell_list_20, entry_cell_list_21, entry_cell_list_22, entry_cell_list_23,
                                entry_cell_list_24, entry_cell_list_25, entry_cell_list_26, entry_cell_list_27, entry_cell_list_28, entry_cell_list_29,
                                entry_cell_list_30, entry_cell_list_31]


        for ec_list in range(0, len(entry_cell_list_list)):
            for entry_cell in range(0, 16):
                entry_cell_list_list[ec_list][entry_cell].grid(row = ec_list + 3, column = entry_cell, padx = 1, pady = .25)


        absence_label = ttk.Label(self.frame_content, text = "Red = Absent")
        absence_label.grid(row=37, column=4, columnspan = 3)
        absence_label.config(background = 'red')
            

        time_label = ttk.Label(self.frame_content, text = "Blue = Non-standard Class Time")
        time_label.grid(row=37, column=9, columnspan = 3)
        time_label.config(background = 'blue')
            

# Methods for teacher search *************************************************************************************************************************
    def search_free_teachers(self):
        teacher_List = Get_Data_From_Excel.get_teacher_list()
        dateList = Get_Data_From_Excel.get_date_list()
        timeList = Get_Data_From_Excel.get_time_list()
        free_teachers_List = teacher_List
        
        date1 = self.entry_combobox1.get()
        time1 = self.entry_combobox2.get()
        date2 = self.entry_combobox3.get()
        time2 = self.entry_combobox4.get()
        date3 = self.entry_combobox5.get()
        time3 = self.entry_combobox6.get()
        date4 = self.entry_combobox7.get()
        time4 = self.entry_combobox8.get()
                        
        date_list = [date1, date2, date3, date4]
        time_list = [time1, time2, time3, time4]
                
        while '' in date_list:
            date_list.remove('')
            
        while '' in time_list:
            time_list.remove('')
   
        for i in range(0, len(date_list)):
                r = dateList.index(date_list[i]) + 2
                c = timeList.index(time_list[i]) + 2
                unavailable_teachers_mini_List = Get_Data_From_Excel.get_unavailable_teachers(r, c)
                for teacher in range(0, len(unavailable_teachers_mini_List)):
                    if unavailable_teachers_mini_List[teacher] in free_teachers_List:
                        free_teachers_List.remove(unavailable_teachers_mini_List[teacher])
                        
        free_teachers_email_list = Get_Data_From_Excel.get_free_teachers_emails(free_teachers_List)
        #insert data found in methods into textbox
        for i in range(0, len(free_teachers_List)):
            self.text_names.insert(0.1, free_teachers_List[i] + '\n')

        for i in range(0, len(free_teachers_email_list)):
            self.text_emails.insert(0.1, free_teachers_email_list[i] + '\n')

    def clear(self):
        self.entry_combobox1.delete(0, 'end')
        self.entry_combobox2.delete(0, 'end')
        self.entry_combobox3.delete(0, 'end')
        self.entry_combobox4.delete(0, 'end')
        self.entry_combobox5.delete(0, 'end')
        self.entry_combobox6.delete(0, 'end')
        self.entry_combobox7.delete(0, 'end')
        self.entry_combobox8.delete(0, 'end')
        self.text_names.delete(0.1, 'end')
        self.text_emails.delete(0.1, 'end')
        
# Methods for emailing schedules to teachers *********************************************************************************************************

    def submit_schedules_for_email(self):
        # This method send each teacher their scheduling information for the next day
        now = datetime.datetime.now()
        schedulingDate = now.day + 1
        schedulingDateRow = now.day + 2 # Note: this is because the time is written in the first row,
                                        # so all dates are moved up one row.
    
        # Get variables from GUI
        sendFrom = self.entry_email.get()
        pwd = self.entry_password.get()
        message = self.text_message.get(1.0, 'end')
        
        # In method of other class -- load spreadsheet and get data
        teacherList = Get_Data_From_Excel.get_teacher_list()
        emailList = Get_Data_From_Excel.get_teacher_emails()

        #log in to email using smtplib module
        smtpObj = smtplib.SMTP('smtp-mail.outlook.com', 587)
        smtpObj.ehlo()
        smtpObj.starttls()
        smtpObj.login(sendFrom, pwd)


        # Loops through multiple sheets in the spreadsheet corresponding to
        # different teachers' schedules
        for s in range(0, len(teacherList)):
            # In method of other class Get variables from spreadsheet cells
            #sendTo = emailList[s]
            sendTo = 'mikalikahihi@yahoo.com' # For testing, uncomment to just send to this one address

            # Get variables from the two lists for data to be sent to teacher
            timeList = Get_Data_From_Excel.get_schedule_times()
            nameList = Get_Data_From_Excel.get_schedule_student_names(s, schedulingDateRow)

           # messageString = '''Subject: Schedule for tomorrow.\n%s''' % (message)    
            messageString = '''Subject: Schedule for tomorrow.\n
                               \r\nHello %s,
                               \r\nThis is your schedule for tomorrow, %s/%s/%s:
                               \r\nMorning: \r\n%s: %s \r\n%s: %s \r\n%s: %s \r\n%s: %s \r\n%s: %s
                               \r\nAfternoon: \r\n%s: %s \r\n%s: %s \r\n%s: %s \r\n%s: %s\r\n%s: %s \r\n%s: %s
                               \r\nEvening: \r\n%s: %s \r\n%s: %s \r\n%s: %s \r\n\r\n%s
                               \r\nHave a great day!''' % (teacherList[s], now.month,
                                                            schedulingDate, now.year,
                                                            timeList[0], nameList[0],
                                                            timeList[1], nameList[1],
                                                            timeList[2], nameList[2],
                                                            timeList[3], nameList[3],
                                                            timeList[4], nameList[4],
                                                            timeList[5], nameList[5],
                                                            timeList[6], nameList[6],
                                                            timeList[7], nameList[7],
                                                            timeList[8], nameList[8],
                                                            timeList[9], nameList[9],
                                                            timeList[10], nameList[10],
                                                            timeList[11], nameList[11],
                                                            timeList[12], nameList[12],
                                                            timeList[13], nameList[13],
                                                            message)
                                                                                                                        
 
            # Send email to teacher -- using smtplib module
            smtpObj.sendmail(sendFrom, sendTo, messageString)

        smtpObj.quit() # logout using smtplib module   
        self.clear()
        messagebox.showinfo(title = "Taipei Language Institute", message = "Messages Sent~")

    def clear2(self):
        self.entry_email.delete(0, 'end')
        self.entry_password.delete(0, 'end')
        self.text_message.delete(1.0, 'end')

# Methods for monthly display of teachers' schedules ************************************************************************************************        
    def search_monthly_schedule(self):
        teacher_search = self.entry_combobox_1.get()
        month = self.entry_combobox_2.get()
        month_info_list = Get_Data_From_Excel.get_current_month_list_and_month_list_values()
        current_month_list = month_info_list[0]
        current_month_list_values = month_info_list[1]
        start_finish_values = Get_Data_From_Excel.get_start_row_finish_row(month, current_month_list, current_month_list_values)
        start_row = start_finish_values[0]
        finish_row = start_finish_values[1]

        # Note: this may be a function in the other class
        class_times = [' ', '7:30-8:20', '8:30-9:20', '9:30-10:20', '10:30-11:20',
                       '11:30-12:20', '12:40-13:30', '13:40-14:30', '14:40-15:30',
                    '15:40-16:30', '16:40-17:30', '17:40-18:30', '18:40-19:30',
                       '19:40-20:30', '20:40-21:30', ' ']

        entry_cell_list = [self.entry_cell_one, self.entry_cell_two, self.entry_cell_three, self.entry_cell_four, self.entry_cell_five,
                            self.entry_cell_six, self.entry_cell_seven, self.entry_cell_eight, self.entry_cell_nine, self.entry_cell_ten,
                            self.entry_cell_eleven, self.entry_cell_twelve, self.entry_cell_thirteen, self.entry_cell_fourteen,
                            self.entry_cell_fifteen, self.entry_cell_sixteen]
        

        entry_cell_list_2 = [self.entry_cell_one_2, self.entry_cell_two_2, self.entry_cell_three_2, self.entry_cell_four_2, self.entry_cell_five_2,
                               self.entry_cell_six_2, self.entry_cell_seven_2, self.entry_cell_eight_2, self.entry_cell_nine_2, self.entry_cell_ten_2,
                               self.entry_cell_eleven_2, self.entry_cell_twelve_2, self.entry_cell_thirteen_2, self.entry_cell_fourteen_2,
                               self.entry_cell_fifteen_2, self.entry_cell_sixteen_2]




        entry_cell_list_3 = [self.entry_cell_one_3, self.entry_cell_two_3, self.entry_cell_three_3, self.entry_cell_four_3, self.entry_cell_five_3,
                               self.entry_cell_six_3, self.entry_cell_seven_3, self.entry_cell_eight_3, self.entry_cell_nine_3, self.entry_cell_ten_3,
                               self.entry_cell_eleven_3, self.entry_cell_twelve_3, self.entry_cell_thirteen_3, self.entry_cell_fourteen_3,
                               self.entry_cell_fifteen_3, self.entry_cell_sixteen_3]




        entry_cell_list_4 = [self.entry_cell_one_4, self.entry_cell_two_4, self.entry_cell_three_4, self.entry_cell_four_4, self.entry_cell_five_4,
                               self.entry_cell_six_4, self.entry_cell_seven_4, self.entry_cell_eight_4, self.entry_cell_nine_4, self.entry_cell_ten_4,
                               self.entry_cell_eleven_4, self.entry_cell_twelve_4, self.entry_cell_thirteen_4, self.entry_cell_fourteen_4,
                               self.entry_cell_fifteen_4, self.entry_cell_sixteen_4]




        entry_cell_list_5 = [self.entry_cell_one_5, self.entry_cell_two_5, self.entry_cell_three_5, self.entry_cell_four_5, self.entry_cell_five_5,
                               self.entry_cell_six_5, self.entry_cell_seven_5, self.entry_cell_eight_5, self.entry_cell_nine_5, self.entry_cell_ten_5,
                               self.entry_cell_eleven_5, self.entry_cell_twelve_5, self.entry_cell_thirteen_5, self.entry_cell_fourteen_5,
                               self.entry_cell_fifteen_5, self.entry_cell_sixteen_5]




        entry_cell_list_6 = [self.entry_cell_one_6, self.entry_cell_two_6, self.entry_cell_three_6, self.entry_cell_four_6, self.entry_cell_five_6,
                               self.entry_cell_six_6, self.entry_cell_seven_6, self.entry_cell_eight_6, self.entry_cell_nine_6, self.entry_cell_ten_6,
                               self.entry_cell_eleven_6, self.entry_cell_twelve_6, self.entry_cell_thirteen_6, self.entry_cell_fourteen_6,
                               self.entry_cell_fifteen_6, self.entry_cell_sixteen_6]




        entry_cell_list_7 = [self.entry_cell_one_7, self.entry_cell_two_7, self.entry_cell_three_7, self.entry_cell_four_7, self.entry_cell_five_7,
                               self.entry_cell_six_7, self.entry_cell_seven_7, self.entry_cell_eight_7, self.entry_cell_nine_7, self.entry_cell_ten_7,
                               self.entry_cell_eleven_7, self.entry_cell_twelve_7, self.entry_cell_thirteen_7, self.entry_cell_fourteen_7,
                               self.entry_cell_fifteen_7, self.entry_cell_sixteen_7]




        entry_cell_list_8 = [self.entry_cell_one_8, self.entry_cell_two_8, self.entry_cell_three_8, self.entry_cell_four_8, self.entry_cell_five_8,
                               self.entry_cell_six_8, self.entry_cell_seven_8, self.entry_cell_eight_8, self.entry_cell_nine_8, self.entry_cell_ten_8,
                               self.entry_cell_eleven_8, self.entry_cell_twelve_8, self.entry_cell_thirteen_8, self.entry_cell_fourteen_8,
                               self.entry_cell_fifteen_8, self.entry_cell_sixteen_8]




        entry_cell_list_9 = [self.entry_cell_one_9, self.entry_cell_two_9, self.entry_cell_three_9, self.entry_cell_four_9, self.entry_cell_five_9,
                               self.entry_cell_six_9, self.entry_cell_seven_9, self.entry_cell_eight_9, self.entry_cell_nine_9, self.entry_cell_ten_9,
                               self.entry_cell_eleven_9, self.entry_cell_twelve_9, self.entry_cell_thirteen_9, self.entry_cell_fourteen_9,
                               self.entry_cell_fifteen_9, self.entry_cell_sixteen_9]




        entry_cell_list_10 = [self.entry_cell_one_10, self.entry_cell_two_10, self.entry_cell_three_10, self.entry_cell_four_10, self.entry_cell_five_10,
                               self.entry_cell_six_10, self.entry_cell_seven_10, self.entry_cell_eight_10, self.entry_cell_nine_10, self.entry_cell_ten_10,
                               self.entry_cell_eleven_10, self.entry_cell_twelve_10, self.entry_cell_thirteen_10, self.entry_cell_fourteen_10,
                               self.entry_cell_fifteen_10, self.entry_cell_sixteen_10]




        entry_cell_list_11 = [self.entry_cell_one_11, self.entry_cell_two_11, self.entry_cell_three_11, self.entry_cell_four_11, self.entry_cell_five_11,
                               self.entry_cell_six_11, self.entry_cell_seven_11, self.entry_cell_eight_11, self.entry_cell_nine_11, self.entry_cell_ten_11,
                               self.entry_cell_eleven_11, self.entry_cell_twelve_11, self.entry_cell_thirteen_11, self.entry_cell_fourteen_11,
                               self.entry_cell_fifteen_11, self.entry_cell_sixteen_11]




        entry_cell_list_12 = [self.entry_cell_one_12, self.entry_cell_two_12, self.entry_cell_three_12, self.entry_cell_four_12, self.entry_cell_five_12,
                               self.entry_cell_six_12, self.entry_cell_seven_12, self.entry_cell_eight_12, self.entry_cell_nine_12, self.entry_cell_ten_12,
                               self.entry_cell_eleven_12, self.entry_cell_twelve_12, self.entry_cell_thirteen_12, self.entry_cell_fourteen_12,
                               self.entry_cell_fifteen_12, self.entry_cell_sixteen_12]




        entry_cell_list_13 = [self.entry_cell_one_13, self.entry_cell_two_13, self.entry_cell_three_13, self.entry_cell_four_13, self.entry_cell_five_13,
                               self.entry_cell_six_13, self.entry_cell_seven_13, self.entry_cell_eight_13, self.entry_cell_nine_13, self.entry_cell_ten_13,
                               self.entry_cell_eleven_13, self.entry_cell_twelve_13, self.entry_cell_thirteen_13, self.entry_cell_fourteen_13,
                               self.entry_cell_fifteen_13, self.entry_cell_sixteen_13]




        entry_cell_list_14 = [self.entry_cell_one_14, self.entry_cell_two_14, self.entry_cell_three_14, self.entry_cell_four_14, self.entry_cell_five_14,
                               self.entry_cell_six_14, self.entry_cell_seven_14, self.entry_cell_eight_14, self.entry_cell_nine_14, self.entry_cell_ten_14,
                               self.entry_cell_eleven_14, self.entry_cell_twelve_14, self.entry_cell_thirteen_14, self.entry_cell_fourteen_14,
                               self.entry_cell_fifteen_14, self.entry_cell_sixteen_14]




        entry_cell_list_15 = [self.entry_cell_one_15, self.entry_cell_two_15, self.entry_cell_three_15, self.entry_cell_four_15, self.entry_cell_five_15,
                               self.entry_cell_six_15, self.entry_cell_seven_15, self.entry_cell_eight_15, self.entry_cell_nine_15, self.entry_cell_ten_15,
                               self.entry_cell_eleven_15, self.entry_cell_twelve_15, self.entry_cell_thirteen_15, self.entry_cell_fourteen_15,
                               self.entry_cell_fifteen_15, self.entry_cell_sixteen_15]




        entry_cell_list_16 = [self.entry_cell_one_16, self.entry_cell_two_16, self.entry_cell_three_16, self.entry_cell_four_16, self.entry_cell_five_16,
                               self.entry_cell_six_16, self.entry_cell_seven_16, self.entry_cell_eight_16, self.entry_cell_nine_16, self.entry_cell_ten_16,
                               self.entry_cell_eleven_16, self.entry_cell_twelve_16, self.entry_cell_thirteen_16, self.entry_cell_fourteen_16,
                               self.entry_cell_fifteen_16, self.entry_cell_sixteen_16]




        entry_cell_list_17 = [self.entry_cell_one_17, self.entry_cell_two_17, self.entry_cell_three_17, self.entry_cell_four_17, self.entry_cell_five_17,
                               self.entry_cell_six_17, self.entry_cell_seven_17, self.entry_cell_eight_17, self.entry_cell_nine_17, self.entry_cell_ten_17,
                               self.entry_cell_eleven_17, self.entry_cell_twelve_17, self.entry_cell_thirteen_17, self.entry_cell_fourteen_17,
                               self.entry_cell_fifteen_17, self.entry_cell_sixteen_17]




        entry_cell_list_18 = [self.entry_cell_one_18, self.entry_cell_two_18, self.entry_cell_three_18, self.entry_cell_four_18, self.entry_cell_five_18,
                               self.entry_cell_six_18, self.entry_cell_seven_18, self.entry_cell_eight_18, self.entry_cell_nine_18, self.entry_cell_ten_18,
                               self.entry_cell_eleven_18, self.entry_cell_twelve_18, self.entry_cell_thirteen_18, self.entry_cell_fourteen_18,
                               self.entry_cell_fifteen_18, self.entry_cell_sixteen_18]




        entry_cell_list_19 = [self.entry_cell_one_19, self.entry_cell_two_19, self.entry_cell_three_19, self.entry_cell_four_19, self.entry_cell_five_19,
                               self.entry_cell_six_19, self.entry_cell_seven_19, self.entry_cell_eight_19, self.entry_cell_nine_19, self.entry_cell_ten_19,
                               self.entry_cell_eleven_19, self.entry_cell_twelve_19, self.entry_cell_thirteen_19, self.entry_cell_fourteen_19,
                               self.entry_cell_fifteen_19, self.entry_cell_sixteen_19]




        entry_cell_list_20 = [self.entry_cell_one_20, self.entry_cell_two_20, self.entry_cell_three_20, self.entry_cell_four_20, self.entry_cell_five_20,
                               self.entry_cell_six_20, self.entry_cell_seven_20, self.entry_cell_eight_20, self.entry_cell_nine_20, self.entry_cell_ten_20,
                               self.entry_cell_eleven_20, self.entry_cell_twelve_20, self.entry_cell_thirteen_20, self.entry_cell_fourteen_20,
                               self.entry_cell_fifteen_20, self.entry_cell_sixteen_20]




        entry_cell_list_21 = [self.entry_cell_one_21, self.entry_cell_two_21, self.entry_cell_three_21, self.entry_cell_four_21, self.entry_cell_five_21,
                               self.entry_cell_six_21, self.entry_cell_seven_21, self.entry_cell_eight_21, self.entry_cell_nine_21, self.entry_cell_ten_21,
                               self.entry_cell_eleven_21, self.entry_cell_twelve_21, self.entry_cell_thirteen_21, self.entry_cell_fourteen_21,
                               self.entry_cell_fifteen_21, self.entry_cell_sixteen_21]




        entry_cell_list_22 = [self.entry_cell_one_22, self.entry_cell_two_22, self.entry_cell_three_22, self.entry_cell_four_22, self.entry_cell_five_22,
                               self.entry_cell_six_22, self.entry_cell_seven_22, self.entry_cell_eight_22, self.entry_cell_nine_22, self.entry_cell_ten_22,
                               self.entry_cell_eleven_22, self.entry_cell_twelve_22, self.entry_cell_thirteen_22, self.entry_cell_fourteen_22,
                               self.entry_cell_fifteen_22, self.entry_cell_sixteen_22]




        entry_cell_list_23 = [self.entry_cell_one_23, self.entry_cell_two_23, self.entry_cell_three_23, self.entry_cell_four_23, self.entry_cell_five_23,
                               self.entry_cell_six_23, self.entry_cell_seven_23, self.entry_cell_eight_23, self.entry_cell_nine_23, self.entry_cell_ten_23,
                               self.entry_cell_eleven_23, self.entry_cell_twelve_23, self.entry_cell_thirteen_23, self.entry_cell_fourteen_23,
                               self.entry_cell_fifteen_23, self.entry_cell_sixteen_23]




        entry_cell_list_24 = [self.entry_cell_one_24, self.entry_cell_two_24, self.entry_cell_three_24, self.entry_cell_four_24, self.entry_cell_five_24,
                               self.entry_cell_six_24, self.entry_cell_seven_24, self.entry_cell_eight_24, self.entry_cell_nine_24, self.entry_cell_ten_24,
                               self.entry_cell_eleven_24, self.entry_cell_twelve_24, self.entry_cell_thirteen_24, self.entry_cell_fourteen_24,
                               self.entry_cell_fifteen_24, self.entry_cell_sixteen_24]




        entry_cell_list_25 = [self.entry_cell_one_25, self.entry_cell_two_25, self.entry_cell_three_25, self.entry_cell_four_25, self.entry_cell_five_25,
                               self.entry_cell_six_25, self.entry_cell_seven_25, self.entry_cell_eight_25, self.entry_cell_nine_25, self.entry_cell_ten_25,
                               self.entry_cell_eleven_25, self.entry_cell_twelve_25, self.entry_cell_thirteen_25, self.entry_cell_fourteen_25,
                               self.entry_cell_fifteen_25, self.entry_cell_sixteen_25]




        entry_cell_list_26 = [self.entry_cell_one_26, self.entry_cell_two_26, self.entry_cell_three_26, self.entry_cell_four_26, self.entry_cell_five_26,
                               self.entry_cell_six_26, self.entry_cell_seven_26, self.entry_cell_eight_26, self.entry_cell_nine_26, self.entry_cell_ten_26,
                               self.entry_cell_eleven_26, self.entry_cell_twelve_26, self.entry_cell_thirteen_26, self.entry_cell_fourteen_26,
                               self.entry_cell_fifteen_26, self.entry_cell_sixteen_26]




        entry_cell_list_27 = [self.entry_cell_one_27, self.entry_cell_two_27, self.entry_cell_three_27, self.entry_cell_four_27, self.entry_cell_five_27,
                               self.entry_cell_six_27, self.entry_cell_seven_27, self.entry_cell_eight_27, self.entry_cell_nine_27, self.entry_cell_ten_27,
                               self.entry_cell_eleven_27, self.entry_cell_twelve_27, self.entry_cell_thirteen_27, self.entry_cell_fourteen_27,
                               self.entry_cell_fifteen_27, self.entry_cell_sixteen_27]




        entry_cell_list_28 = [self.entry_cell_one_28, self.entry_cell_two_28, self.entry_cell_three_28, self.entry_cell_four_28, self.entry_cell_five_28,
                               self.entry_cell_six_28, self.entry_cell_seven_28, self.entry_cell_eight_28, self.entry_cell_nine_28, self.entry_cell_ten_28,
                               self.entry_cell_eleven_28, self.entry_cell_twelve_28, self.entry_cell_thirteen_28, self.entry_cell_fourteen_28,
                               self.entry_cell_fifteen_28, self.entry_cell_sixteen_28]




        entry_cell_list_29 = [self.entry_cell_one_29, self.entry_cell_two_29, self.entry_cell_three_29, self.entry_cell_four_29, self.entry_cell_five_29,
                               self.entry_cell_six_29, self.entry_cell_seven_29, self.entry_cell_eight_29, self.entry_cell_nine_29, self.entry_cell_ten_29,
                               self.entry_cell_eleven_29, self.entry_cell_twelve_29, self.entry_cell_thirteen_29, self.entry_cell_fourteen_29,
                               self.entry_cell_fifteen_29, self.entry_cell_sixteen_29]




        entry_cell_list_30 = [self.entry_cell_one_30, self.entry_cell_two_30, self.entry_cell_three_30, self.entry_cell_four_30, self.entry_cell_five_30,
                               self.entry_cell_six_30, self.entry_cell_seven_30, self.entry_cell_eight_30, self.entry_cell_nine_30, self.entry_cell_ten_30,
                               self.entry_cell_eleven_30, self.entry_cell_twelve_30, self.entry_cell_thirteen_30, self.entry_cell_fourteen_30,
                               self.entry_cell_fifteen_30, self.entry_cell_sixteen_30]




        entry_cell_list_31 = [self.entry_cell_one_31, self.entry_cell_two_31, self.entry_cell_three_31, self.entry_cell_four_31, self.entry_cell_five_31,
                               self.entry_cell_six_31, self.entry_cell_seven_31, self.entry_cell_eight_31, self.entry_cell_nine_31, self.entry_cell_ten_31,
                               self.entry_cell_eleven_31, self.entry_cell_twelve_31, self.entry_cell_thirteen_31, self.entry_cell_fourteen_31,
                               self.entry_cell_fifteen_31, self.entry_cell_sixteen_31]


        entry_cell_list_list = [entry_cell_list, entry_cell_list_2, entry_cell_list_3, entry_cell_list_4, entry_cell_list_5, entry_cell_list_6,
                                    entry_cell_list_7, entry_cell_list_8, entry_cell_list_9, entry_cell_list_10, entry_cell_list_11, entry_cell_list_11,
                                    entry_cell_list_12, entry_cell_list_13, entry_cell_list_14, entry_cell_list_15, entry_cell_list_16, entry_cell_list_17,
                                    entry_cell_list_18, entry_cell_list_19, entry_cell_list_20, entry_cell_list_21, entry_cell_list_22, entry_cell_list_23,
                                    entry_cell_list_24, entry_cell_list_25, entry_cell_list_26, entry_cell_list_27, entry_cell_list_28, entry_cell_list_29,
                                    entry_cell_list_30, entry_cell_list_31]

        for ec_list in range(0, len(entry_cell_list_list)):
            for item in range(0, len(entry_cell_list)):
                entry_cell_list_list[ec_list][item].delete(0, 'end')


        box_value_list = class_times

            
        for item in range(0, len(box_value_list)):
            entry_cell_list[item].insert(0, box_value_list[item])

        work_book = Get_Data_From_Excel.open_workbook()
        entry_cell_list_number = 1   
        row_number = start_row

        while row_number < finish_row:
            box_value_list = Get_Data_From_Excel.get_students_scheduled_for_classes(work_book, teacher_search, row_number)
            for item in range(0, len(box_value_list)):
                entry_cell_list_list[entry_cell_list_number][item].insert(0, box_value_list[item])
                if 'ABSENT' in box_value_list[item]:
                    entry_cell_list_list[entry_cell_list_number][item].config(foreground='red')
                elif 'Absent' in box_value_list[item]:
                    entry_cell_list_list[entry_cell_list_number][item].config(foreground='red')
                elif 'absent' in box_value_list[item]:
                    entry_cell_list_list[entry_cell_list_number][item].config(foreground='red')
                elif '(' in box_value_list[item]:
                    entry_cell_list_list[entry_cell_list_number][item].config(foreground='dark blue')
            entry_cell_list_number += 1
            row_number += 1
            time.sleep(0.01)
        
    def clear3(self):
        self.entry_combobox_1.delete(0, 'end')
        self.entry_combobox_2.delete(0, 'end')

        entry_cell_list = [self.entry_cell_one, self.entry_cell_two, self.entry_cell_three, self.entry_cell_four, self.entry_cell_five,
                            self.entry_cell_six, self.entry_cell_seven, self.entry_cell_eight, self.entry_cell_nine, self.entry_cell_ten,
                            self.entry_cell_eleven, self.entry_cell_twelve, self.entry_cell_thirteen, self.entry_cell_fourteen,
                            self.entry_cell_fifteen, self.entry_cell_sixteen]
        

        entry_cell_list_2 = [self.entry_cell_one_2, self.entry_cell_two_2, self.entry_cell_three_2, self.entry_cell_four_2, self.entry_cell_five_2,
                               self.entry_cell_six_2, self.entry_cell_seven_2, self.entry_cell_eight_2, self.entry_cell_nine_2, self.entry_cell_ten_2,
                               self.entry_cell_eleven_2, self.entry_cell_twelve_2, self.entry_cell_thirteen_2, self.entry_cell_fourteen_2,
                               self.entry_cell_fifteen_2, self.entry_cell_sixteen_2]




        entry_cell_list_3 = [self.entry_cell_one_3, self.entry_cell_two_3, self.entry_cell_three_3, self.entry_cell_four_3, self.entry_cell_five_3,
                               self.entry_cell_six_3, self.entry_cell_seven_3, self.entry_cell_eight_3, self.entry_cell_nine_3, self.entry_cell_ten_3,
                               self.entry_cell_eleven_3, self.entry_cell_twelve_3, self.entry_cell_thirteen_3, self.entry_cell_fourteen_3,
                               self.entry_cell_fifteen_3, self.entry_cell_sixteen_3]




        entry_cell_list_4 = [self.entry_cell_one_4, self.entry_cell_two_4, self.entry_cell_three_4, self.entry_cell_four_4, self.entry_cell_five_4,
                               self.entry_cell_six_4, self.entry_cell_seven_4, self.entry_cell_eight_4, self.entry_cell_nine_4, self.entry_cell_ten_4,
                               self.entry_cell_eleven_4, self.entry_cell_twelve_4, self.entry_cell_thirteen_4, self.entry_cell_fourteen_4,
                               self.entry_cell_fifteen_4, self.entry_cell_sixteen_4]




        entry_cell_list_5 = [self.entry_cell_one_5, self.entry_cell_two_5, self.entry_cell_three_5, self.entry_cell_four_5, self.entry_cell_five_5,
                               self.entry_cell_six_5, self.entry_cell_seven_5, self.entry_cell_eight_5, self.entry_cell_nine_5, self.entry_cell_ten_5,
                               self.entry_cell_eleven_5, self.entry_cell_twelve_5, self.entry_cell_thirteen_5, self.entry_cell_fourteen_5,
                               self.entry_cell_fifteen_5, self.entry_cell_sixteen_5]




        entry_cell_list_6 = [self.entry_cell_one_6, self.entry_cell_two_6, self.entry_cell_three_6, self.entry_cell_four_6, self.entry_cell_five_6,
                               self.entry_cell_six_6, self.entry_cell_seven_6, self.entry_cell_eight_6, self.entry_cell_nine_6, self.entry_cell_ten_6,
                               self.entry_cell_eleven_6, self.entry_cell_twelve_6, self.entry_cell_thirteen_6, self.entry_cell_fourteen_6,
                               self.entry_cell_fifteen_6, self.entry_cell_sixteen_6]




        entry_cell_list_7 = [self.entry_cell_one_7, self.entry_cell_two_7, self.entry_cell_three_7, self.entry_cell_four_7, self.entry_cell_five_7,
                               self.entry_cell_six_7, self.entry_cell_seven_7, self.entry_cell_eight_7, self.entry_cell_nine_7, self.entry_cell_ten_7,
                               self.entry_cell_eleven_7, self.entry_cell_twelve_7, self.entry_cell_thirteen_7, self.entry_cell_fourteen_7,
                               self.entry_cell_fifteen_7, self.entry_cell_sixteen_7]




        entry_cell_list_8 = [self.entry_cell_one_8, self.entry_cell_two_8, self.entry_cell_three_8, self.entry_cell_four_8, self.entry_cell_five_8,
                               self.entry_cell_six_8, self.entry_cell_seven_8, self.entry_cell_eight_8, self.entry_cell_nine_8, self.entry_cell_ten_8,
                               self.entry_cell_eleven_8, self.entry_cell_twelve_8, self.entry_cell_thirteen_8, self.entry_cell_fourteen_8,
                               self.entry_cell_fifteen_8, self.entry_cell_sixteen_8]




        entry_cell_list_9 = [self.entry_cell_one_9, self.entry_cell_two_9, self.entry_cell_three_9, self.entry_cell_four_9, self.entry_cell_five_9,
                               self.entry_cell_six_9, self.entry_cell_seven_9, self.entry_cell_eight_9, self.entry_cell_nine_9, self.entry_cell_ten_9,
                               self.entry_cell_eleven_9, self.entry_cell_twelve_9, self.entry_cell_thirteen_9, self.entry_cell_fourteen_9,
                               self.entry_cell_fifteen_9, self.entry_cell_sixteen_9]




        entry_cell_list_10 = [self.entry_cell_one_10, self.entry_cell_two_10, self.entry_cell_three_10, self.entry_cell_four_10, self.entry_cell_five_10,
                               self.entry_cell_six_10, self.entry_cell_seven_10, self.entry_cell_eight_10, self.entry_cell_nine_10, self.entry_cell_ten_10,
                               self.entry_cell_eleven_10, self.entry_cell_twelve_10, self.entry_cell_thirteen_10, self.entry_cell_fourteen_10,
                               self.entry_cell_fifteen_10, self.entry_cell_sixteen_10]




        entry_cell_list_11 = [self.entry_cell_one_11, self.entry_cell_two_11, self.entry_cell_three_11, self.entry_cell_four_11, self.entry_cell_five_11,
                               self.entry_cell_six_11, self.entry_cell_seven_11, self.entry_cell_eight_11, self.entry_cell_nine_11, self.entry_cell_ten_11,
                               self.entry_cell_eleven_11, self.entry_cell_twelve_11, self.entry_cell_thirteen_11, self.entry_cell_fourteen_11,
                               self.entry_cell_fifteen_11, self.entry_cell_sixteen_11]




        entry_cell_list_12 = [self.entry_cell_one_12, self.entry_cell_two_12, self.entry_cell_three_12, self.entry_cell_four_12, self.entry_cell_five_12,
                               self.entry_cell_six_12, self.entry_cell_seven_12, self.entry_cell_eight_12, self.entry_cell_nine_12, self.entry_cell_ten_12,
                               self.entry_cell_eleven_12, self.entry_cell_twelve_12, self.entry_cell_thirteen_12, self.entry_cell_fourteen_12,
                               self.entry_cell_fifteen_12, self.entry_cell_sixteen_12]




        entry_cell_list_13 = [self.entry_cell_one_13, self.entry_cell_two_13, self.entry_cell_three_13, self.entry_cell_four_13, self.entry_cell_five_13,
                               self.entry_cell_six_13, self.entry_cell_seven_13, self.entry_cell_eight_13, self.entry_cell_nine_13, self.entry_cell_ten_13,
                               self.entry_cell_eleven_13, self.entry_cell_twelve_13, self.entry_cell_thirteen_13, self.entry_cell_fourteen_13,
                               self.entry_cell_fifteen_13, self.entry_cell_sixteen_13]




        entry_cell_list_14 = [self.entry_cell_one_14, self.entry_cell_two_14, self.entry_cell_three_14, self.entry_cell_four_14, self.entry_cell_five_14,
                               self.entry_cell_six_14, self.entry_cell_seven_14, self.entry_cell_eight_14, self.entry_cell_nine_14, self.entry_cell_ten_14,
                               self.entry_cell_eleven_14, self.entry_cell_twelve_14, self.entry_cell_thirteen_14, self.entry_cell_fourteen_14,
                               self.entry_cell_fifteen_14, self.entry_cell_sixteen_14]




        entry_cell_list_15 = [self.entry_cell_one_15, self.entry_cell_two_15, self.entry_cell_three_15, self.entry_cell_four_15, self.entry_cell_five_15,
                               self.entry_cell_six_15, self.entry_cell_seven_15, self.entry_cell_eight_15, self.entry_cell_nine_15, self.entry_cell_ten_15,
                               self.entry_cell_eleven_15, self.entry_cell_twelve_15, self.entry_cell_thirteen_15, self.entry_cell_fourteen_15,
                               self.entry_cell_fifteen_15, self.entry_cell_sixteen_15]




        entry_cell_list_16 = [self.entry_cell_one_16, self.entry_cell_two_16, self.entry_cell_three_16, self.entry_cell_four_16, self.entry_cell_five_16,
                               self.entry_cell_six_16, self.entry_cell_seven_16, self.entry_cell_eight_16, self.entry_cell_nine_16, self.entry_cell_ten_16,
                               self.entry_cell_eleven_16, self.entry_cell_twelve_16, self.entry_cell_thirteen_16, self.entry_cell_fourteen_16,
                               self.entry_cell_fifteen_16, self.entry_cell_sixteen_16]




        entry_cell_list_17 = [self.entry_cell_one_17, self.entry_cell_two_17, self.entry_cell_three_17, self.entry_cell_four_17, self.entry_cell_five_17,
                               self.entry_cell_six_17, self.entry_cell_seven_17, self.entry_cell_eight_17, self.entry_cell_nine_17, self.entry_cell_ten_17,
                               self.entry_cell_eleven_17, self.entry_cell_twelve_17, self.entry_cell_thirteen_17, self.entry_cell_fourteen_17,
                               self.entry_cell_fifteen_17, self.entry_cell_sixteen_17]




        entry_cell_list_18 = [self.entry_cell_one_18, self.entry_cell_two_18, self.entry_cell_three_18, self.entry_cell_four_18, self.entry_cell_five_18,
                               self.entry_cell_six_18, self.entry_cell_seven_18, self.entry_cell_eight_18, self.entry_cell_nine_18, self.entry_cell_ten_18,
                               self.entry_cell_eleven_18, self.entry_cell_twelve_18, self.entry_cell_thirteen_18, self.entry_cell_fourteen_18,
                               self.entry_cell_fifteen_18, self.entry_cell_sixteen_18]




        entry_cell_list_19 = [self.entry_cell_one_19, self.entry_cell_two_19, self.entry_cell_three_19, self.entry_cell_four_19, self.entry_cell_five_19,
                               self.entry_cell_six_19, self.entry_cell_seven_19, self.entry_cell_eight_19, self.entry_cell_nine_19, self.entry_cell_ten_19,
                               self.entry_cell_eleven_19, self.entry_cell_twelve_19, self.entry_cell_thirteen_19, self.entry_cell_fourteen_19,
                               self.entry_cell_fifteen_19, self.entry_cell_sixteen_19]




        entry_cell_list_20 = [self.entry_cell_one_20, self.entry_cell_two_20, self.entry_cell_three_20, self.entry_cell_four_20, self.entry_cell_five_20,
                               self.entry_cell_six_20, self.entry_cell_seven_20, self.entry_cell_eight_20, self.entry_cell_nine_20, self.entry_cell_ten_20,
                               self.entry_cell_eleven_20, self.entry_cell_twelve_20, self.entry_cell_thirteen_20, self.entry_cell_fourteen_20,
                               self.entry_cell_fifteen_20, self.entry_cell_sixteen_20]




        entry_cell_list_21 = [self.entry_cell_one_21, self.entry_cell_two_21, self.entry_cell_three_21, self.entry_cell_four_21, self.entry_cell_five_21,
                               self.entry_cell_six_21, self.entry_cell_seven_21, self.entry_cell_eight_21, self.entry_cell_nine_21, self.entry_cell_ten_21,
                               self.entry_cell_eleven_21, self.entry_cell_twelve_21, self.entry_cell_thirteen_21, self.entry_cell_fourteen_21,
                               self.entry_cell_fifteen_21, self.entry_cell_sixteen_21]




        entry_cell_list_22 = [self.entry_cell_one_22, self.entry_cell_two_22, self.entry_cell_three_22, self.entry_cell_four_22, self.entry_cell_five_22,
                               self.entry_cell_six_22, self.entry_cell_seven_22, self.entry_cell_eight_22, self.entry_cell_nine_22, self.entry_cell_ten_22,
                               self.entry_cell_eleven_22, self.entry_cell_twelve_22, self.entry_cell_thirteen_22, self.entry_cell_fourteen_22,
                               self.entry_cell_fifteen_22, self.entry_cell_sixteen_22]




        entry_cell_list_23 = [self.entry_cell_one_23, self.entry_cell_two_23, self.entry_cell_three_23, self.entry_cell_four_23, self.entry_cell_five_23,
                               self.entry_cell_six_23, self.entry_cell_seven_23, self.entry_cell_eight_23, self.entry_cell_nine_23, self.entry_cell_ten_23,
                               self.entry_cell_eleven_23, self.entry_cell_twelve_23, self.entry_cell_thirteen_23, self.entry_cell_fourteen_23,
                               self.entry_cell_fifteen_23, self.entry_cell_sixteen_23]




        entry_cell_list_24 = [self.entry_cell_one_24, self.entry_cell_two_24, self.entry_cell_three_24, self.entry_cell_four_24, self.entry_cell_five_24,
                               self.entry_cell_six_24, self.entry_cell_seven_24, self.entry_cell_eight_24, self.entry_cell_nine_24, self.entry_cell_ten_24,
                               self.entry_cell_eleven_24, self.entry_cell_twelve_24, self.entry_cell_thirteen_24, self.entry_cell_fourteen_24,
                               self.entry_cell_fifteen_24, self.entry_cell_sixteen_24]




        entry_cell_list_25 = [self.entry_cell_one_25, self.entry_cell_two_25, self.entry_cell_three_25, self.entry_cell_four_25, self.entry_cell_five_25,
                               self.entry_cell_six_25, self.entry_cell_seven_25, self.entry_cell_eight_25, self.entry_cell_nine_25, self.entry_cell_ten_25,
                               self.entry_cell_eleven_25, self.entry_cell_twelve_25, self.entry_cell_thirteen_25, self.entry_cell_fourteen_25,
                               self.entry_cell_fifteen_25, self.entry_cell_sixteen_25]




        entry_cell_list_26 = [self.entry_cell_one_26, self.entry_cell_two_26, self.entry_cell_three_26, self.entry_cell_four_26, self.entry_cell_five_26,
                               self.entry_cell_six_26, self.entry_cell_seven_26, self.entry_cell_eight_26, self.entry_cell_nine_26, self.entry_cell_ten_26,
                               self.entry_cell_eleven_26, self.entry_cell_twelve_26, self.entry_cell_thirteen_26, self.entry_cell_fourteen_26,
                               self.entry_cell_fifteen_26, self.entry_cell_sixteen_26]




        entry_cell_list_27 = [self.entry_cell_one_27, self.entry_cell_two_27, self.entry_cell_three_27, self.entry_cell_four_27, self.entry_cell_five_27,
                               self.entry_cell_six_27, self.entry_cell_seven_27, self.entry_cell_eight_27, self.entry_cell_nine_27, self.entry_cell_ten_27,
                               self.entry_cell_eleven_27, self.entry_cell_twelve_27, self.entry_cell_thirteen_27, self.entry_cell_fourteen_27,
                               self.entry_cell_fifteen_27, self.entry_cell_sixteen_27]




        entry_cell_list_28 = [self.entry_cell_one_28, self.entry_cell_two_28, self.entry_cell_three_28, self.entry_cell_four_28, self.entry_cell_five_28,
                               self.entry_cell_six_28, self.entry_cell_seven_28, self.entry_cell_eight_28, self.entry_cell_nine_28, self.entry_cell_ten_28,
                               self.entry_cell_eleven_28, self.entry_cell_twelve_28, self.entry_cell_thirteen_28, self.entry_cell_fourteen_28,
                               self.entry_cell_fifteen_28, self.entry_cell_sixteen_28]




        entry_cell_list_29 = [self.entry_cell_one_29, self.entry_cell_two_29, self.entry_cell_three_29, self.entry_cell_four_29, self.entry_cell_five_29,
                               self.entry_cell_six_29, self.entry_cell_seven_29, self.entry_cell_eight_29, self.entry_cell_nine_29, self.entry_cell_ten_29,
                               self.entry_cell_eleven_29, self.entry_cell_twelve_29, self.entry_cell_thirteen_29, self.entry_cell_fourteen_29,
                               self.entry_cell_fifteen_29, self.entry_cell_sixteen_29]




        entry_cell_list_30 = [self.entry_cell_one_30, self.entry_cell_two_30, self.entry_cell_three_30, self.entry_cell_four_30, self.entry_cell_five_30,
                               self.entry_cell_six_30, self.entry_cell_seven_30, self.entry_cell_eight_30, self.entry_cell_nine_30, self.entry_cell_ten_30,
                               self.entry_cell_eleven_30, self.entry_cell_twelve_30, self.entry_cell_thirteen_30, self.entry_cell_fourteen_30,
                               self.entry_cell_fifteen_30, self.entry_cell_sixteen_30]




        entry_cell_list_31 = [self.entry_cell_one_31, self.entry_cell_two_31, self.entry_cell_three_31, self.entry_cell_four_31, self.entry_cell_five_31,
                               self.entry_cell_six_31, self.entry_cell_seven_31, self.entry_cell_eight_31, self.entry_cell_nine_31, self.entry_cell_ten_31,
                               self.entry_cell_eleven_31, self.entry_cell_twelve_31, self.entry_cell_thirteen_31, self.entry_cell_fourteen_31,
                               self.entry_cell_fifteen_31, self.entry_cell_sixteen_31]


        entry_cell_list_list = [entry_cell_list, entry_cell_list_2, entry_cell_list_3, entry_cell_list_4, entry_cell_list_5, entry_cell_list_6,
                                    entry_cell_list_7, entry_cell_list_8, entry_cell_list_9, entry_cell_list_10, entry_cell_list_11, entry_cell_list_11,
                                    entry_cell_list_12, entry_cell_list_13, entry_cell_list_14, entry_cell_list_15, entry_cell_list_16, entry_cell_list_17,
                                    entry_cell_list_18, entry_cell_list_19, entry_cell_list_20, entry_cell_list_21, entry_cell_list_22, entry_cell_list_23,
                                    entry_cell_list_24, entry_cell_list_25, entry_cell_list_26, entry_cell_list_27, entry_cell_list_28, entry_cell_list_29,
                                    entry_cell_list_30, entry_cell_list_31]


        for ec_list in range(0, len(entry_cell_list_list)):
            for item in range(0, len(entry_cell_list)):
                entry_cell_list_list[ec_list][item].delete(0, 'end')
   




#======================================================================================================================================================     
# Mainloop

def main():
   
    root = Tk()
    feedback = GUI_Input_Output(root)
    root.mainloop()

if __name__ == "__main__": main()
    
